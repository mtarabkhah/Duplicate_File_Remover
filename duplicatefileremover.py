import os
import hashlib
import shutil
import math
from PIL import Image
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from send2trash import send2trash

def get_image_orientation(image_path):
    try:
        with Image.open(image_path) as image:
            exif = image._getexif()
            if exif is not None and 274 in exif:
                return exif[274]
    except (IOError, OSError):
        pass
    return None

def find_duplicate_files(directories, extensions):
    file_hashes = {}
    duplicated_files = []
    unique_files_size = 0
    duplicated_versions_size = 0
    all_files_size = 0
    all_files_count = 0
    unique_files_count = 0
    duplicated_files_count = 0

    for directory in directories:
        for root, _, files in os.walk(directory):
            for filename in files:
                if filename.lower().endswith(tuple(extensions)):
                    file_path = os.path.join(root, filename)

                    file_size = os.path.getsize(file_path)
                    
                    all_files_count += 1
                    all_files_size += file_size

                    with open(file_path, 'rb') as f:
                        file_hash = hashlib.md5(f.read()).hexdigest()
                        # Please note that this code uses the MD5 hash function to compare file content, which is generally fast but may have a small chance of collision 
                        # (i.e., two different files having the same hash). If you require a more secure comparison, you can replace hashlib.md5 with hashlib.sha256. 
                        # However, keep in mind that the SHA-256 hash function is slower than MD5.

                    # # Check image orientation for rotated versions
                    # orientation = get_image_orientation(file_path)
                    # if orientation is None:
                    #     continue
                        
                    # if orientation in [3, 6, 8]:
                    #     file_hash = hashlib.md5(f'{file_hash}-{orientation}'.encode('utf-8')).hexdigest()
                        
                    if file_hash in file_hashes:
                        # file_hashes[file_hash][1] should be equal to file_size otherwise they are different files
                        # It's a good idea to check the file sizes before adding them in the duplicated list                         
                        duplicated_versions_size += file_size
                        duplicated_files_count += 1
                        if file_hashes[file_hash][2] is None:
                            ind = unique_files_count
                            file_hashes[file_hash][2] = ind
                            duplicated_files.append([[file_hashes[file_hash][0]], 0, file_hash])
                            unique_files_count +=1
                            unique_files_size += file_size
                        else:
                            ind = file_hashes[file_hash][2]
                        duplicated_files[ind][0].append(file_path)
                        duplicated_files[ind][1] += file_size
                    else:
                        file_hashes[file_hash] = [file_path,file_size,None]
    
    return all_files_count, all_files_size, file_hashes, duplicated_files, unique_files_count, unique_files_size, duplicated_files_count, duplicated_versions_size


def adjust_column_width(sheet):
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = max_length
        sheet.column_dimensions[column_letter].width = adjusted_width


def generate_report(all_files_count, all_files_size, file_hashes, duplicated_files, unique_files_count, unique_files_size, duplicated_files_count, duplicated_versions_size):
    
    duplicated_files = sorted(duplicated_files, key=lambda x: x[1], reverse=True)

    print("Total number of files:", all_files_count)
    print("Total size of files:", convert_size(all_files_size))
    print("Total number of unique files:", unique_files_count)
    print("Total size of unique files:", convert_size(unique_files_size))
    print("Total number of duplicated files:", duplicated_files_count)
    print("Total size of duplicated versions:", convert_size(duplicated_versions_size))

    # Create a new workbook
    workbook = Workbook()

    # Create the first sheet (original files and sizes + number and size of their duplicated versions)
    sheet1 = workbook.active
    sheet1.title = "Original Files"
    sheet1.append(["Original File", "Original File Size", "Duplicated files' count", "Duplicated files' size"])
    for i in range(unique_files_count):
        hash = duplicated_files[i][2]
        sheet1.append([file_hashes[hash][0], convert_size(file_hashes[hash][1]), len(duplicated_files[i][0])-1, convert_size(duplicated_files[i][1])])
    adjust_column_width(sheet1)
    

    # Create the second sheet (path for duplicated files groups)
    sheet2 = workbook.create_sheet(title="Duplicated Files")
    sheet2.append(["Group","File Path", "File size"])
    for i in range(unique_files_count):
        sheet2.append([i+1, "Size of duplicated versions", convert_size(duplicated_files[i][1])])
        for j in range(len(duplicated_files[i][0])):
            sheet2.append(["", duplicated_files[i][0][j], ""])
        sheet2.append(["", "", ""])
    adjust_column_width(sheet2)

    # Save the workbook
    workbook.save("duplicate_files_report.xlsx")
    
    
def delete_files(path_list, delete_permenantly):
    for file_path in path_list:
        if delete_permenantly:
            os.remove(file_path) # permanently deleting the files 
        else:
            send2trash(file_path) # sending the file to the recycle bin/trash
            # Please note that the behavior of the send2trash function may vary depending on the operating system and configuration. 
            # It is recommended to test this functionality on your specific system to ensure it works as expected.
        print("Deleted:", file_path)
        print("------------------------")
    

def move_duplicates(path_list, destination_folder):
    for file_path in path_list:
        filename = os.path.basename(file_path)
        destination_path = os.path.join(destination_folder, filename)

        # Move the duplicate file to the destination folder
        shutil.move(file_path, destination_path)

        print("Moved:", filename)
        print("To:", destination_path)
        print("------------------------")


def convert_size(size_bytes):
    # Convert bytes to a human-readable format
    if size_bytes == 0:
        return "0B"
    size_name = ("B", "KB", "MB", "GB", "TB", "PB", "EB", "ZB", "YB")
    i = int(math.floor(math.log(size_bytes, 1024)))
    p = math.pow(1024, i)
    size = round(size_bytes / p, 2)
    return f"{size} {size_name[i]}"


# Example directories and extensions
current_directory = os.getcwd()
directories = [current_directory+'\Test photos - Copy\F1', current_directory+'\Test photos - Copy\F2',
               current_directory+'\Test photos - Copy\F3']
extensions = ['.jpg', '.jpeg', '.png', '.mp4', '.mov', '.avi']


# Call the function to find duplicate files
all_files_count, all_files_size, file_hashes, duplicated_files, unique_files_count, unique_files_size, duplicated_files_count, duplicated_versions_size = find_duplicate_files(directories, extensions)


# Call the function to Generate report
generate_report(all_files_count, all_files_size, file_hashes, duplicated_files, unique_files_count, unique_files_size, duplicated_files_count, duplicated_versions_size)


# Provide the destination folder to move the duplicates or leave it empty for deleting
destination_folder = current_directory+'\Test photos - Copy\Duplicated'
delete_flag = True
delete_permenantly = False

if len(duplicated_files) > 0:
    # Provide the index of the main file to keep (default is 0, the first duplicate)
    main_file_index = 0

    path_list = []
    for file_paths, _, _ in duplicated_files:
        main_file = file_paths[main_file_index]
        for file_path in file_paths:
            if file_path in main_file:
                continue
            else:
                path_list.append(file_path)


    if delete_flag:
        # Delete the duplicated files instead of moving
        delete_files(path_list, delete_permenantly)
        # duplicated_files_to_delete = [file_path for _, file_path, _ in duplicated_files]
        # delete_files(duplicated_files_to_delete)
    else:
        move_duplicates(path_list, destination_folder)        
